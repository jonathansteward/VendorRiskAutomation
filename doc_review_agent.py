import json
import logging
import os
import re
from datetime import date
from typing import Callable

logger = logging.getLogger(__name__)

import openai
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from vendor_risk_engine import (
    ATTESTATION_MULTIPLIERS,
    calculate_control_strength,
    calculate_loss_magnitude,
    calculate_residual_risk,
    CONTROL_CATEGORIES,
)

def _get_client():
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise EnvironmentError(
            "OPENAI_API_KEY environment variable is not set. "
            "Get your key at https://platform.openai.com/api-keys"
        )
    return openai.OpenAI(api_key=api_key)

# ---------------------------------------------------------------------------
# Human-readable display names for each control key
# ---------------------------------------------------------------------------
CONTROL_DISPLAY_NAMES = {
    # Access
    "multi_factor_authentication":            "Multi-Factor Authentication (MFA)",
    "privileged_access_management":           "Privileged Access Management (PAM)",
    "least_privilege_rbac":                   "Least Privilege / RBAC",
    "identity_lifecycle_management":          "Identity Lifecycle Management",
    "periodic_access_reviews":               "Periodic Access Reviews",
    "single_sign_on_central_identity":        "Single Sign-On / Central Identity",
    "session_management_timeout_controls":    "Session Management / Timeout Controls",
    "authentication_logging_monitoring":      "Authentication Logging & Monitoring",
    "credential_storage_security":            "Credential Storage Security",
    "account_lockout_brute_force_protection": "Account Lockout / Brute Force Protection",
    # Data Security
    "encryption_at_rest":          "Encryption at Rest",
    "encryption_in_transit":       "Encryption in Transit",
    "data_access_restrictions":    "Data Access Restrictions",
    "data_classification_program": "Data Classification Program",
    "database_security_controls":  "Database Security Controls",
    "data_loss_prevention":        "Data Loss Prevention (DLP)",
    "key_management_security":     "Key Management Security",
    "secure_backup_protection":    "Secure Backup Protection",
    "data_retention_policies":     "Data Retention Policies",
    "data_integrity_validation":   "Data Integrity Validation",
    # Integration
    "api_authentication":             "API Authentication",
    "secure_credential_storage":      "Secure Credential Storage",
    "input_validation_sanitization":  "Input Validation / Sanitization",
    "api_authorization_controls":     "API Authorization Controls",
    "rate_limiting_abuse_protection": "Rate Limiting / Abuse Protection",
    "integration_logging":            "Integration Logging",
    "transport_security_tls":         "Transport Security (TLS)",
    "api_gateway_security":           "API Gateway Security",
    "vendor_integration_reviews":     "Vendor Integration Reviews",
    "service_account_restrictions":   "Service Account Restrictions",
    # AI
    "training_data_governance":    "Training Data Governance",
    "model_access_control":        "Model Access Control",
    "model_validation_testing":    "Model Validation / Testing",
    "output_monitoring":           "Output Monitoring",
    "prompt_injection_protection": "Prompt Injection Protection",
    "human_oversight_review":      "Human Oversight / Review",
    "model_version_control":       "Model Version Control",
    "model_security_monitoring":   "Model Security Monitoring",
    "model_input_sanitization":    "Model Input Sanitization",
    "third_party_ai_risk_review":  "Third-Party AI Risk Review",
    # Availability
    "system_redundancy_failover":   "System Redundancy / Failover",
    "backup_restore_capability":    "Backup & Restore Capability",
    "disaster_recovery_plan":       "Disaster Recovery Plan",
    "infrastructure_monitoring":    "Infrastructure Monitoring",
    "capacity_planning":            "Capacity Planning",
    "incident_response_procedures": "Incident Response Procedures",
    "network_resilience":           "Network Resilience",
    "patch_management":             "Patch Management",
    "load_balancing":               "Load Balancing",
    "service_recovery_testing":     "Service Recovery Testing",
    # Governance
    "security_policy_framework":          "Security Policy Framework",
    "risk_management_program":            "Risk Management Program",
    "third_party_risk_management":        "Third-Party Risk Management",
    "security_awareness_training":        "Security Awareness Training",
    "compliance_certifications":          "Compliance Certifications",
    "security_incident_reporting":        "Security Incident Reporting",
    "vulnerability_management_program":   "Vulnerability Management Program",
    "change_management":                  "Change Management",
    "audit_program":                      "Audit Program",
    "security_leadership_governance":     "Security Leadership / Governance",
}

CATEGORY_DISPLAY_NAMES = {
    "access":        "Access Risk Controls",
    "data_security": "Data Security Risk Controls",
    "integration":   "Integration Risk Controls",
    "ai":            "AI Risk Controls",
    "availability":  "Availability Risk Controls",
    "governance":    "Governance Risk Controls",
}

SCORE_TO_STATUS = {1.0: "Implemented", 0.5: "Partially Implemented", 0.0: "Not Implemented"}

# ---------------------------------------------------------------------------
# System prompt
# ---------------------------------------------------------------------------
SYSTEM_PROMPT = """You are a senior cybersecurity analyst performing vendor risk assessments
using the FAIR risk model. You evaluate vendor security documentation against control categories.

For each control, you must assess:
1. Implementation score: 1.0 = Fully Implemented, 0.5 = Partially Implemented, 0.0 = Not Implemented
2. Where in the document the evidence was found (section name, page number, or "Not found")
3. The exact language or paraphrase from the document that satisfies the control intent
   (or "No evidence found" if absent)
4. Whether follow-up is required from the vendor (true/false)
5. What specific information needs follow-up (or "" if none)

Scoring guidance:
- Attestation/audit reports (SOC 2 Type II, ISO 27001, pen test): if the auditor confirmed the
  control operated effectively, score 1.0 and note the finding in 'where_found'
  (e.g., "SOC 2 Type II — CC6.1, tested effective"). Score 0.5 if noted with exceptions.
- Policy/procedure documents that merely state a control exists: score 0.5 unless implementation
  evidence is provided (configuration details, screenshots, training records, etc.) — then 1.0.
- A control not mentioned anywhere in the documentation: score 0.0.

Be precise and evidence-based. Quote or closely paraphrase the document."""


def _build_user_prompt(vendor_name: str, applicable_categories: set = None) -> str:
    """Build the structured review prompt; excludes AI controls for non-AI-enabled vendors."""
    cats = applicable_categories or set(CONTROL_CATEGORIES.keys())
    # Preserve canonical category order
    ordered_cats = [c for c in CONTROL_CATEGORIES if c in cats]
    n_controls   = sum(len(CONTROL_CATEGORIES[c]) for c in ordered_cats)
    ai_note      = "" if "ai" in cats else \
        "\nNote: AI Risk Controls are excluded — this vendor's product/service is not AI-enabled.\n"

    # Build the JSON scaffold dynamically so AI block is omitted when not applicable
    blank = '{"score": 0, "where_found": "", "language": "", "follow_up_required": false, "follow_up_info": ""}'
    cat_blocks = []
    for cat in ordered_cats:
        ctrl_lines = ",\n      ".join(
            f'"{k}": {blank}'
            for k in CONTROL_CATEGORIES[cat]
        )
        cat_blocks.append(f'    "{cat}": {{\n      {ctrl_lines}\n    }}')
    details_json = ",\n".join(cat_blocks)

    return (
        f'Review the security documentation for {vendor_name} against {n_controls} controls '
        f'across {len(ordered_cats)} categories below.{ai_note}\n\n'
        f'Return ONLY valid JSON (no markdown fences) using exactly this structure.\n'
        f'Use only 0, 0.5, or 1.0 for scores. Every key must be present.\n\n'
        f'{{\n'
        f'  "vendor": "{vendor_name}",\n'
        f'  "control_details": {{\n{details_json}\n  }},\n'
        f'  "gaps": [],\n'
        f'  "overall_summary": ""\n'
        f'}}'
    )


# ---------------------------------------------------------------------------
# Core review function
# ---------------------------------------------------------------------------

SUPPORTED_EXTENSIONS = {".pdf", ".txt", ".md", ".docx"}

# Valid score values the LLM is allowed to return (A08 — insecure deserialization)
VALID_SCORES = {0, 0.5, 1.0}

# All expected control keys per category — used for schema validation (A08)
# Built from CONTROL_CATEGORIES to stay in sync with the engine automatically.
EXPECTED_CONTROLS: dict[str, set[str]] = {
    cat: set(controls.keys())
    for cat, controls in CONTROL_CATEGORIES.items()
}


def _validate_llm_response(review: dict, applicable_categories: set = None) -> None:
    """
    Validate the structure and values of the LLM JSON response (A08).
    Only validates categories in applicable_categories (AI omitted for non-AI vendors).
    Raises ValueError with a descriptive message on any violation.
    """
    if not isinstance(review, dict):
        raise ValueError("LLM response is not a JSON object.")

    if "control_details" not in review:
        raise ValueError("LLM response missing 'control_details' key.")

    cats_to_check = applicable_categories or set(EXPECTED_CONTROLS.keys())

    details = review["control_details"]
    for category, expected_keys in EXPECTED_CONTROLS.items():
        if category not in cats_to_check:
            continue  # skip non-applicable categories (e.g. AI for non-AI vendors)
        if category not in details:
            raise ValueError(f"LLM response missing category '{category}'.")

        cat_data = details[category]
        for ctrl_key in expected_keys:
            if ctrl_key not in cat_data:
                raise ValueError(f"LLM response missing control '{ctrl_key}' in '{category}'.")

            entry = cat_data[ctrl_key]
            if not isinstance(entry, dict):
                raise ValueError(f"Control entry '{ctrl_key}' in '{category}' is not an object.")

            raw_score = entry.get("score")
            # Coerce to float to handle integer 0/1 from LLM
            try:
                score = float(raw_score)
            except (TypeError, ValueError):
                raise ValueError(f"Invalid score '{raw_score}' for '{ctrl_key}' in '{category}'.")

            if score not in VALID_SCORES:
                # Snap to nearest valid value rather than rejecting outright
                snapped = min(VALID_SCORES, key=lambda v: abs(v - score))
                logger.warning("Score %.2f for '%s/%s' snapped to %.1f", score, category, ctrl_key, snapped)
                entry["score"] = snapped
            else:
                entry["score"] = score

            if not isinstance(entry.get("follow_up_required"), bool):
                entry["follow_up_required"] = bool(entry.get("follow_up_required", False))

    # Sanitize free-text fields to prevent stored XSS if rendered in UI (A03)
    for category in details:
        for ctrl_key in details[category]:
            entry = details[category][ctrl_key]
            for field in ("where_found", "language", "follow_up_info"):
                if field in entry:
                    entry[field] = _sanitize_text(str(entry[field]))

    if "gaps" in review:
        review["gaps"] = [_sanitize_text(str(g)) for g in review["gaps"] if g]

    if "overall_summary" in review:
        review["overall_summary"] = _sanitize_text(str(review.get("overall_summary", "")))


def _sanitize_text(text: str) -> str:
    """Strip HTML/script tags to prevent stored XSS (A03).

    html.unescape() is applied first so that encoded variants like
    &lt;script&gt; are normalised before the tag-stripping regex runs.
    """
    import html
    return re.sub(r"<[^>]+>", "", html.unescape(text)).strip()[:2000]


def _collect_files(doc_source: str) -> list[str]:
    """
    Accept either a single file path or a folder path.
    Returns a sorted list of supported document file paths.
    """
    if os.path.isfile(doc_source):
        return [doc_source]

    if os.path.isdir(doc_source):
        files = sorted(
            os.path.join(doc_source, f)
            for f in os.listdir(doc_source)
            if os.path.splitext(f)[1].lower() in SUPPORTED_EXTENSIONS
        )
        if not files:
            raise ValueError(
                f"No supported documents found in '{doc_source}'. "
                f"Supported types: {', '.join(SUPPORTED_EXTENSIONS)}"
            )
        return files

    raise ValueError(f"'{doc_source}' is not a valid file or directory.")


def _extract_text(file_path: str) -> str:
    """Extract plain text from PDF, DOCX, TXT, or MD files."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        return "\n".join(page.extract_text() or "" for page in reader.pages)

    if ext == ".docx":
        from docx import Document
        doc = Document(file_path)
        return "\n".join(p.text for p in doc.paragraphs)

    # .txt / .md
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def _build_questionnaire_data(applicable_categories: set = None) -> dict:
    """Build the questionnaire structure to send to the frontend.
    Excludes categories not in applicable_categories (e.g. AI for non-AI vendors)."""
    cats = applicable_categories or set(CONTROL_CATEGORIES.keys())
    return {
        category: {
            "label": CATEGORY_DISPLAY_NAMES.get(category, category),
            "controls": [
                {
                    "key": key,
                    "name": CONTROL_DISPLAY_NAMES.get(key, key.replace("_", " ").title()),
                    "weight": weight,
                }
                for key, weight in controls.items()
            ],
        }
        for category, controls in CONTROL_CATEGORIES.items()
        if category in cats
    }


def _build_from_questionnaire(
    vendor_name: str,
    questionnaire_answers: dict,
    emit: Callable[[str], None],
    applicable_categories: set = None,
) -> tuple[dict, list[str], str]:
    """Build control_details, gaps, and overall_summary from questionnaire answers."""
    cats = applicable_categories or set(CONTROL_CATEGORIES.keys())
    control_details: dict = {}
    gaps: list[str] = []

    for category, controls in CONTROL_CATEGORIES.items():
        if category not in cats:
            continue
        control_details[category] = {}
        cat_answers = questionnaire_answers.get(category, {})

        for control_key, weight in controls.items():
            raw = cat_answers.get(control_key, 0)
            try:
                level = float(raw)
            except (TypeError, ValueError):
                level = 0.0
            level = min(VALID_SCORES, key=lambda v: abs(v - level))

            control_details[category][control_key] = {
                "score": level,
                "where_found": "Questionnaire" if level > 0 else "Not reported",
                "language": SCORE_TO_STATUS.get(level, "Not Implemented"),
                "follow_up_required": level < 1.0,
                "follow_up_info": (
                    "Self-reported — supporting documentation recommended"
                    if 0 < level < 1.0
                    else ("Not implemented — action required" if level == 0 else "")
                ),
            }

            if level < 0.5 and weight >= 0.10:
                name = CONTROL_DISPLAY_NAMES.get(control_key, control_key)
                cat_label = CATEGORY_DISPLAY_NAMES.get(category, category)
                gaps.append(
                    f"{cat_label}: {name} — not implemented (weight {weight:.0%})"
                )

    total       = sum(len(c) for c in control_details.values())
    implemented = sum(1 for c in control_details.values() for v in c.values() if v["score"] >= 1.0)
    partial     = sum(1 for c in control_details.values() for v in c.values() if v["score"] == 0.5)

    overall_summary = (
        f"Questionnaire-based assessment for {vendor_name}: "
        f"{implemented}/{total} controls fully implemented, {partial} partially implemented. "
        "Validate self-reported control status with supporting documentation."
    )
    emit(
        f"Questionnaire processed: {implemented} implemented, {partial} partial, "
        f"{total - implemented - partial} not implemented out of {total} controls."
    )
    return control_details, gaps, overall_summary


def generate_mitigation_recommendations(
    category_effectiveness: dict,
    control_details: dict,
    n_recommendations: int = 7,
) -> list[dict]:
    """
    Return the top N controls to implement for maximum risk reduction.

    Priority = weight × implementation_gap × (1 − category_score).
    Only includes categories that are Ineffective or Partially Effective.
    """
    recommendations: list[dict] = []

    for category, effectiveness in category_effectiveness.items():
        if effectiveness.get("rating") == "Effective":
            continue

        cat_score  = effectiveness.get("score", 0.0)
        weights    = CONTROL_CATEGORIES.get(category, {})
        cat_ctrls  = control_details.get(category, {})

        for control_key, weight in weights.items():
            detail      = cat_ctrls.get(control_key, {})
            current_impl = float(detail.get("score", 0.0)) if isinstance(detail, dict) else float(detail)

            if current_impl >= 1.0:
                continue

            gap      = 1.0 - current_impl
            priority = weight * gap * (1.0 - cat_score)

            recommendations.append({
                "category":       category,
                "category_label": CATEGORY_DISPLAY_NAMES.get(category, category),
                "control_key":    control_key,
                "control_name":   CONTROL_DISPLAY_NAMES.get(
                    control_key, control_key.replace("_", " ").title()
                ),
                "current_status": "Not Implemented" if current_impl == 0 else "Partially Implemented",
                "current_score":  current_impl,
                "weight":         weight,
                "priority_score": round(priority, 4),
            })

    recommendations.sort(key=lambda x: x["priority_score"], reverse=True)
    return recommendations[:n_recommendations]


def review_security_documentation(
    doc_source: str | None,
    vendor_name: str,
    contact_frequency: float,
    probability_of_action: float,
    threat_capability: float,
    resistance_strength: float,
    loss_magnitude: float,
    company_revenue: float = 0.0,
    questionnaire_answers: dict | None = None,
    progress_callback: Callable[[str], None] | None = None,
    is_ai_enabled: bool = True,
    has_attestation: bool = False,
    loss_magnitude_components: dict | None = None,
    distributions: dict | None = None,
) -> dict:
    """
    Review vendor security documentation (or questionnaire answers) and compute FAIR risk.

    Args:
        doc_source:               Path to a file or folder, or None if no documents.
        questionnaire_answers:    Pre-scored controls from the UI questionnaire.
                                  When provided, skips LLM doc review entirely.
        progress_callback:        Optional callable(msg) for real-time progress updates.
        is_ai_enabled:            Whether the vendor's product/service uses AI/ML features.
                                  When False, AI Risk Controls are excluded from scoring
                                  and do not contribute to the residual risk calculation.
        loss_magnitude_components: Dict of structured loss magnitude inputs for
                                  calculate_loss_magnitude(). When provided, replaces the
                                  loss_magnitude scalar with a component-based total.
                                  Keys: breach_notification_cost, regulatory_fine_exposure,
                                  incident_response_cost, downtime_cost_per_hour,
                                  estimated_downtime_hours, reputation_damage_pct,
                                  annual_revenue.
        distributions:            Dict of (min, likely, max) tuples for Monte Carlo simulation.
                                  Keys: contact_frequency, probability_of_action,
                                  threat_capability, resistance_strength, loss_magnitude.
                                  When provided, P10/P50/P90 distribution output is included.
    """
    def emit(msg: str):
        if progress_callback:
            progress_callback(msg)
        else:
            print(msg)

    # --- Resolve applicable categories (exclude AI for non-AI vendors) ---
    applicable_categories = set(CONTROL_CATEGORIES.keys())
    if not is_ai_enabled:
        applicable_categories.discard("ai")
        emit("AI Risk Controls excluded — vendor product/service is not AI-enabled.")

    # --- Attestation multiplier — set by user checkbox, not LLM auto-detection ---
    # ×1.0 when third-party attestation is confirmed — controls accepted at full face value.
    # ×0.85 when no attestation — self-reported controls are discounted 15% for lack of
    # independent verification (policy doc stating a control ≠ auditor confirming it works).
    attestation_multiplier = 1.0 if has_attestation else 0.85
    if has_attestation:
        emit("Third-party attestation confirmed — controls accepted at full value (×1.0).")
    else:
        emit("No third-party attestation — applying ×0.85 skepticism discount to control scores.")

    # --- Resolve loss magnitude (structured components take precedence) ---
    if loss_magnitude_components:
        lm_result = calculate_loss_magnitude(**loss_magnitude_components)
        loss_magnitude = lm_result["total"]
        lm_dist = lm_result["distribution"]
        lm_source = "Structured component calculation (FAIR primary + secondary loss)"
        emit(f"Structured loss magnitude: ${loss_magnitude:,.0f} "
             f"(range ${lm_dist['min']:,.0f}–${lm_dist['max']:,.0f})")
        # Add loss magnitude distribution for Monte Carlo
        if distributions is None:
            distributions = {}
        distributions["loss_magnitude"] = (lm_dist["min"], lm_dist["likely"], lm_dist["max"])
    else:
        lm_result = None
        lm_source = "Product/service revenue estimate"

    # --- Determine control scores source ---
    if questionnaire_answers is not None:
        # Use questionnaire answers directly — no LLM review needed
        control_details, gaps, overall_summary = _build_from_questionnaire(
            vendor_name, questionnaire_answers, emit, applicable_categories
        )
    elif doc_source:
        # LLM document review path
        file_paths = _collect_files(doc_source)
        emit(f"Found {len(file_paths)} document(s) for {vendor_name}:")
        for fp in file_paths:
            emit(f"  • {os.path.basename(fp)}")

        doc_sections = []
        for fp in file_paths:
            emit(f"Extracting text from {os.path.basename(fp)}...")
            text = _extract_text(fp)
            # Use only basename to avoid leaking filesystem paths (A02)
            doc_sections.append(
                f"=== DOCUMENT: {os.path.basename(fp)} ===\n{text}\n=== END ==="
            )

        combined_docs = "\n\n".join(doc_sections)
        emit(f"All {len(file_paths)} document(s) extracted. Reviewing with GPT-4o...")

        user_message = (
            f"The following is the complete security documentation package provided by {vendor_name}. "
            f"Review all documents together as a unified evidence set — a control may be evidenced "
            f"across multiple files.\n\n"
            f"{combined_docs}\n\n"
            f"{_build_user_prompt(vendor_name, applicable_categories)}"
        )

        response = _get_client().chat.completions.create(
            model="gpt-4o",
            max_tokens=16000,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": user_message},
            ],
            response_format={"type": "json_object"},
        )

        raw = response.choices[0].message.content.strip()
        raw = raw.removeprefix("```json").removeprefix("```").removesuffix("```").strip()

        try:
            review = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise ValueError(f"LLM returned invalid JSON: {exc}") from exc

        # Validate and sanitize all LLM output before use (A08, A03)
        _validate_llm_response(review, applicable_categories)

        control_details  = review["control_details"]
        gaps             = review.get("gaps", [])
        overall_summary  = review.get("overall_summary", "")
    else:
        # No documents and no questionnaire — signal the caller to request questionnaire
        return {
            "questionnaire_required": True,
            "questionnaire_data": _build_questionnaire_data(applicable_categories),
            "is_ai_enabled": is_ai_enabled,
        }

    # --- Calculate control effectiveness scores ---
    control_scores_flat = {
        category: {k: v["score"] for k, v in controls.items()}
        for category, controls in control_details.items()
    }

    emit("Calculating control effectiveness scores...")
    category_results: dict        = {}
    control_scores_for_risk: dict = {}

    for category, controls in control_scores_flat.items():
        result = calculate_control_strength(category, controls, attestation_multiplier)
        category_results[category]        = result
        control_scores_for_risk[category] = result["score"]
        attest_note = " [attested]" if result.get("attestation_applied") else ""
        emit(f"  {category:<15} score={result['score']:.4f}  rating={result['rating']}{attest_note}")

    # --- Build input provenance for audit trail (#7) ---
    input_sources = {
        "contact_frequency_source":     "Threat intelligence (DBIR / CISA / FBI IC3 / MITRE ATT&CK)",
        "probability_of_action_source": "Threat intelligence (Verizon DBIR actor motivation data)",
        "threat_capability_source":     "Threat intelligence (MITRE ATT&CK technique prevalence)",
        "resistance_strength_source":   "Threat intelligence (DBIR sector baseline security posture)",
        "loss_magnitude_source":        lm_source,
        "control_scores_source":        "LLM document review (GPT-4o)" if doc_source else "User questionnaire",
        "attestation_multiplier":       attestation_multiplier,
        "has_attestation":              has_attestation,
        "methodology":                  "FAIR (Open FAIR Body of Knowledge v2.0) + Monte Carlo triangular distributions",
        "reduction_ceiling_reference":  "NIST SP 800-30 Rev.1 / IBM 2024 Cost of a Data Breach Report",
        "governance_split_reference":   "FAIR-CAM v1.0 / SANS administrative control classification",
    }

    emit("Calculating FAIR risk scores...")
    risk = calculate_residual_risk(
        contact_frequency=contact_frequency,
        probability_of_action=probability_of_action,
        threat_capability=threat_capability,
        resistance_strength=resistance_strength,
        loss_magnitude=loss_magnitude,
        control_scores=control_scores_for_risk,
        company_revenue=company_revenue,
        applicable_categories=applicable_categories,
        input_sources=input_sources,
        distributions=distributions,
    )

    # --- Mitigation recommendations ---
    recommendations = generate_mitigation_recommendations(category_results, control_details)
    if recommendations:
        emit(f"Top mitigation: {recommendations[0]['control_name']} ({recommendations[0]['category_label']})")

    return {
        "vendor":                    vendor_name,
        "assessment_date":           date.today().isoformat(),
        "control_details":           control_details,
        "control_effectiveness":     category_results,
        "gaps":                      gaps,
        "overall_summary":           overall_summary,
        "risk":                      risk,
        "fair_inputs": {
            "contact_frequency":     contact_frequency,
            "probability_of_action": probability_of_action,
            "threat_capability":     threat_capability,
            "resistance_strength":   resistance_strength,
            "product_revenue":       loss_magnitude,
            "company_revenue":       company_revenue,
            "loss_magnitude_components": lm_result["components"] if lm_result else None,
        },
        "is_ai_enabled":             is_ai_enabled,
        "applicable_categories":     sorted(applicable_categories),
        "has_attestation":           has_attestation,
        "attestation_multiplier":    attestation_multiplier,
        "input_provenance":          input_sources,
        "mitigation_recommendations": recommendations,
    }


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

def export_control_review_xlsx(assessment: dict, output_path: str) -> None:
    """Export the per-control review to an XLSX file."""

    # Colour fills
    fill_green  = PatternFill("solid", fgColor="C6EFCE")
    fill_yellow = PatternFill("solid", fgColor="FFEB9C")
    fill_red    = PatternFill("solid", fgColor="FFC7CE")
    fill_header = PatternFill("solid", fgColor="1F3864")
    fill_cat    = PatternFill("solid", fgColor="2F5496")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wrap = Alignment(wrap_text=True, vertical="top")
    center_wrap = Alignment(wrap_text=True, vertical="top", horizontal="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control Review"

    # Header row
    headers = [
        "Control Category",
        "Control Description",
        "Implementation Status",
        "Document Where Found",
        "Language That Met Intent of Control",
        "Follow-Up Required",
        "What Information Needs Follow-Up",
    ]
    col_widths = [28, 38, 22, 28, 55, 16, 45]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = fill_header
        cell.alignment = center_wrap
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 32

    current_row = 2
    for category_key, controls in assessment["control_details"].items():
        category_name = CATEGORY_DISPLAY_NAMES.get(category_key, category_key)

        # Category separator row
        cat_cell = ws.cell(row=current_row, column=1, value=category_name)
        cat_cell.font = Font(bold=True, color="FFFFFF", size=10)
        cat_cell.fill = fill_cat
        cat_cell.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(headers))
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for control_key, detail in controls.items():
            score = detail.get("score", 0)
            status = SCORE_TO_STATUS.get(score, "Not Implemented")
            follow_up = "Yes" if detail.get("follow_up_required") else "No"

            row_data = [
                category_name,
                CONTROL_DISPLAY_NAMES.get(control_key, control_key),
                status,
                detail.get("where_found", ""),
                detail.get("language", ""),
                follow_up,
                detail.get("follow_up_info", ""),
            ]

            if score == 1.0:
                status_fill = fill_green
            elif score == 0.5:
                status_fill = fill_yellow
            else:
                status_fill = fill_red

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = border
                cell.alignment = wrap if col_idx != 3 else center_wrap
                cell.font = Font(size=10)
                if col_idx == 3:
                    cell.fill = status_fill

            ws.row_dimensions[current_row].height = 60
            current_row += 1

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"Control review saved: {output_path}")


# ---------------------------------------------------------------------------
# PDF report
# ---------------------------------------------------------------------------

def export_risk_report_pdf(assessment: dict, output_path: str) -> None:
    """Export the FAIR risk report to a PDF with Summary, Details, and Conclusion."""

    vendor = assessment["vendor"]
    risk = assessment["risk"]
    gaps = assessment["gaps"]
    effectiveness = assessment["control_effectiveness"]
    fair_inputs = assessment["fair_inputs"]
    assessment_date = assessment.get("assessment_date", date.today().isoformat())

    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )

    styles = getSampleStyleSheet()

    # Custom styles
    style_title = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=20,
        textColor=colors.HexColor("#1F3864"),
        spaceAfter=4,
    )
    style_subtitle = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=11,
        textColor=colors.HexColor("#2F5496"),
        spaceAfter=16,
    )
    style_h1 = ParagraphStyle(
        "H1",
        parent=styles["Heading1"],
        fontSize=14,
        textColor=colors.HexColor("#1F3864"),
        spaceBefore=18,
        spaceAfter=6,
        borderPad=4,
    )
    style_h2 = ParagraphStyle(
        "H2",
        parent=styles["Heading2"],
        fontSize=11,
        textColor=colors.HexColor("#2F5496"),
        spaceBefore=12,
        spaceAfter=4,
    )
    style_body = ParagraphStyle(
        "Body",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        spaceAfter=6,
    )
    style_bullet = ParagraphStyle(
        "Bullet",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        leftIndent=16,
        spaceAfter=3,
    )

    def hr():
        return HRFlowable(width="100%", thickness=1,
                          color=colors.HexColor("#2F5496"), spaceAfter=8)

    def risk_color(inherent, residual):
        pct = (residual / inherent * 100) if inherent else 100
        if pct <= 30:
            return colors.HexColor("#C6EFCE"), colors.HexColor("#375623")
        elif pct <= 60:
            return colors.HexColor("#FFEB9C"), colors.HexColor("#7D6608")
        return colors.HexColor("#FFC7CE"), colors.HexColor("#9C0006")

    def rating_color(rating):
        if rating == "Effective":
            return colors.HexColor("#C6EFCE")
        elif rating == "Partially Effective":
            return colors.HexColor("#FFEB9C")
        return colors.HexColor("#FFC7CE")

    story = []

    # ---- Cover / Title ----
    story.append(Paragraph("Vendor Security Risk Assessment", style_title))
    story.append(Paragraph(
        f"{vendor}  &nbsp;|&nbsp;  Assessment Date: {assessment_date}", style_subtitle
    ))
    story.append(hr())

    # ================================================================
    # SECTION 1 — SUMMARY
    # ================================================================
    story.append(Paragraph("1. Summary", style_h1))
    story.append(hr())

    story.append(Paragraph(
        f"This report presents the results of a FAIR-based vendor security risk assessment "
        f"conducted for <b>{vendor}</b>. The assessment evaluated 60 security controls across "
        f"six risk categories using the vendor's security documentation as the primary evidence source.",
        style_body
    ))

    if assessment.get("overall_summary"):
        story.append(Paragraph(assessment["overall_summary"], style_body))

    story.append(Spacer(1, 10))
    story.append(Paragraph("FAIR Risk Inputs", style_h2))

    inputs_data = [
        ["Parameter", "Value"],
        ["Contact Frequency (events/year)", f"{fair_inputs['contact_frequency']}"],
        ["Probability of Action", f"{fair_inputs['probability_of_action']:.0%}"],
        ["Threat Capability", f"{fair_inputs['threat_capability']:.0%}"],
        ["Baseline Resistance Strength", f"{fair_inputs['resistance_strength']:.0%}"],
        ["Product/Service Revenue (Loss Magnitude)", f"${fair_inputs.get('product_revenue', fair_inputs.get('loss_magnitude', 0)):,.0f}"],
    ]
    if fair_inputs.get("company_revenue"):
        inputs_data.append(["Company Annual Revenue", f"${fair_inputs['company_revenue']:,.0f}"])
    inputs_table = Table(inputs_data, colWidths=[3.5 * inch, 2.5 * inch])
    inputs_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF1F7")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(inputs_table)

    story.append(Spacer(1, 14))
    story.append(Paragraph("Risk Scores", style_h2))

    residual_bg, residual_fg = risk_color(risk["inherent_risk"], risk["residual_risk"])
    reduction_pct = (
        risk["risk_reduction"] / risk["inherent_risk"] * 100
        if risk["inherent_risk"] else 0
    )

    inherent_rating = risk.get("inherent_rating", {})
    residual_rating = risk.get("residual_rating", {})

    def _are_label(rating_dict: dict) -> str:
        parts = [f"ARE {rating_dict.get('are_pct', 0):.1f}% of product revenue"]
        if "are_pct_company" in rating_dict:
            parts.append(f"{rating_dict['are_pct_company']:.2f}% of company revenue")
        return "  |  ".join(parts)

    risk_data = [
        ["", "Annual Cost Exposure", "Rating"],
        [
            "Inherent Risk (no controls)",
            f"${risk['inherent_risk']:,.2f}",
            f"{inherent_rating.get('rating', '')}  —  {_are_label(inherent_rating)}",
        ],
        [
            "Residual Risk (controls applied)",
            f"${risk['residual_risk']:,.2f}",
            f"{residual_rating.get('rating', '')}  —  {_are_label(residual_rating)}",
        ],
        [
            "Risk Reduction",
            f"${risk['risk_reduction']:,.2f}",
            f"{reduction_pct:.1f}% reduction",
        ],
    ]
    risk_table = Table(risk_data, colWidths=[2.2 * inch, 1.6 * inch, 3.2 * inch])
    risk_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 10),
        ("FONTNAME",   (0, 1), (0, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 2), (-1, 2), residual_bg),
        ("TEXTCOLOR",  (0, 2), (-1, 2), residual_fg),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(risk_table)

    story.append(Spacer(1, 14))
    story.append(Paragraph("Control Effectiveness Summary", style_h2))

    eff_data = [["Control Category", "Score", "Rating"]]
    for cat_key, result in effectiveness.items():
        eff_data.append([
            CATEGORY_DISPLAY_NAMES.get(cat_key, cat_key),
            f"{result['score']:.2f}",
            result["rating"],
        ])

    eff_table = Table(eff_data, colWidths=[3.5 * inch, 1.2 * inch, 2.3 * inch])
    eff_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for row_idx, (cat_key, result) in enumerate(effectiveness.items(), start=1):
        bg = rating_color(result["rating"])
        eff_style.append(("BACKGROUND", (2, row_idx), (2, row_idx), bg))

    eff_table.setStyle(TableStyle(eff_style))
    story.append(eff_table)

    # ================================================================
    # SECTION 2 — VENDOR SECURITY ASSESSMENT DETAILS
    # ================================================================
    story.append(Paragraph("2. Vendor Security Assessment Details", style_h1))
    story.append(hr())

    story.append(Paragraph(
        "The following section details the control effectiveness findings for each risk category, "
        "including composite scores and identified control gaps.",
        style_body,
    ))

    for cat_key, result in effectiveness.items():
        cat_name = CATEGORY_DISPLAY_NAMES.get(cat_key, cat_key)
        story.append(Paragraph(cat_name, style_h2))

        story.append(Paragraph(
            f"<b>Overall Score:</b> {result['score']:.2f} &nbsp;&nbsp; "
            f"<b>Rating:</b> {result['rating']}",
            style_body,
        ))

        # Per-control score table for this category
        controls = assessment["control_details"].get(cat_key, {})
        ctrl_data = [["Control", "Status", "Follow-Up"]]
        for ctrl_key, detail in controls.items():
            score = detail.get("score", 0)
            ctrl_data.append([
                CONTROL_DISPLAY_NAMES.get(ctrl_key, ctrl_key),
                SCORE_TO_STATUS.get(score, "Not Implemented"),
                "Yes" if detail.get("follow_up_required") else "No",
            ])

        ctrl_table = Table(ctrl_data, colWidths=[3.8 * inch, 2 * inch, 1.2 * inch])
        ctrl_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF1F7")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        for row_idx, (ctrl_key, detail) in enumerate(controls.items(), start=1):
            score = detail.get("score", 0)
            bg = (fill_green  := colors.HexColor("#C6EFCE")) if score == 1.0 else \
                 (colors.HexColor("#FFEB9C") if score == 0.5 else colors.HexColor("#FFC7CE"))
            ctrl_style.append(("BACKGROUND", (1, row_idx), (1, row_idx), bg))

        ctrl_table.setStyle(TableStyle(ctrl_style))
        story.append(ctrl_table)
        story.append(Spacer(1, 6))

        if result.get("unscored_controls"):
            story.append(Paragraph(
                f"<i>Controls with no evidence (scored 0): "
                f"{', '.join(result['unscored_controls'])}</i>",
                style_bullet,
            ))

    # ================================================================
    # SECTION 3 — CONCLUSION
    # ================================================================
    story.append(Paragraph("3. Conclusion", style_h1))
    story.append(hr())

    # Key findings
    effective_cats   = [CATEGORY_DISPLAY_NAMES.get(k) for k, v in effectiveness.items() if v["rating"] == "Effective"]
    partial_cats     = [CATEGORY_DISPLAY_NAMES.get(k) for k, v in effectiveness.items() if v["rating"] == "Partially Effective"]
    ineffective_cats = [CATEGORY_DISPLAY_NAMES.get(k) for k, v in effectiveness.items() if v["rating"] == "Ineffective"]

    story.append(Paragraph("Key Findings", style_h2))

    if effective_cats:
        story.append(Paragraph(
            f"<b>Effective control areas:</b> {', '.join(effective_cats)}.", style_body
        ))
    if partial_cats:
        story.append(Paragraph(
            f"<b>Partially effective areas requiring improvement:</b> {', '.join(partial_cats)}.", style_body
        ))
    if ineffective_cats:
        story.append(Paragraph(
            f"<b>Ineffective areas with significant control gaps:</b> {', '.join(ineffective_cats)}.", style_body
        ))

    story.append(Paragraph(
        f"The application of assessed controls reduces annualized risk exposure from "
        f"<b>${risk['inherent_risk']:,.2f}</b> (inherent) to "
        f"<b>${risk['residual_risk']:,.2f}</b> (residual), "
        f"representing a <b>{reduction_pct:.1f}% reduction</b> in expected annual loss.",
        style_body,
    ))

    # Critical gaps
    if gaps:
        story.append(Paragraph("Identified Gaps", style_h2))
        for gap in gaps:
            story.append(Paragraph(f"• {gap}", style_bullet))

    # Follow-up items
    follow_up_items = []
    for cat_key, controls in assessment["control_details"].items():
        for ctrl_key, detail in controls.items():
            if detail.get("follow_up_required") and detail.get("follow_up_info"):
                follow_up_items.append((
                    CATEGORY_DISPLAY_NAMES.get(cat_key, cat_key),
                    CONTROL_DISPLAY_NAMES.get(ctrl_key, ctrl_key),
                    detail["follow_up_info"],
                ))

    if follow_up_items:
        story.append(Paragraph("Follow-Up Items Required from Vendor", style_h2))
        fu_data = [["Category", "Control", "Information Needed"]]
        for cat_name, ctrl_name, info in follow_up_items:
            fu_data.append([cat_name, ctrl_name, info])

        fu_table = Table(fu_data, colWidths=[1.8 * inch, 2.2 * inch, 3 * inch])
        fu_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF2CC")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("WORDWRAP", (0, 0), (-1, -1), True),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(fu_table)

    story.append(Spacer(1, 14))
    story.append(Paragraph(
        "This assessment is based solely on the documentation provided by the vendor. "
        "Findings should be validated through direct vendor engagement and, where appropriate, "
        "independent verification or on-site review.",
        ParagraphStyle("Disclaimer", parent=style_body,
                       textColor=colors.HexColor("#595959"), fontSize=9),
    ))

    doc.build(story)
    print(f"PDF report saved: {output_path}")


# ---------------------------------------------------------------------------
# Console summary (kept for quick review)
# ---------------------------------------------------------------------------

def print_report(assessment: dict) -> None:
    vendor = assessment["vendor"]
    risk = assessment["risk"]

    print("\n" + "=" * 60)
    print(f"  VENDOR RISK ASSESSMENT: {vendor}")
    print("=" * 60)

    print("\nCONTROL EFFECTIVENESS")
    print("-" * 60)
    for category, result in assessment["control_effectiveness"].items():
        score = result["score"]
        bar = "#" * int(score * 20)
        print(f"  {category:<15} {score:.2f} [{bar:<20}] {result['rating']}")

    print("\nIDENTIFIED GAPS")
    print("-" * 60)
    for gap in assessment["gaps"]:
        print(f"  • {gap}")

    print("\nFAIR RISK SCORES")
    print("-" * 60)
    print(f"  Inherent Risk:  ${risk['inherent_risk']:>14,.2f} / year")
    print(f"  Residual Risk:  ${risk['residual_risk']:>14,.2f} / year")
    print(f"  Risk Reduction: ${risk['risk_reduction']:>14,.2f} / year")
    print("=" * 60 + "\n")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        print("Usage: python doc_review_agent.py <path_to_pdf> <vendor_name>")
        print("Example: python doc_review_agent.py vendor_soc2.pdf 'Acme Corp'")
        sys.exit(1)

    doc_path = sys.argv[1]
    vendor_name = sys.argv[2]

    # FAIR inputs — adjust per assessment
    CONTACT_FREQUENCY     = 12
    PROBABILITY_OF_ACTION = 0.3
    THREAT_CAPABILITY     = 0.7
    RESISTANCE_STRENGTH   = 0.3
    LOSS_MAGNITUDE        = 1_000_000

    assessment = review_security_documentation(
        doc_path=doc_path,
        vendor_name=vendor_name,
        contact_frequency=CONTACT_FREQUENCY,
        probability_of_action=PROBABILITY_OF_ACTION,
        threat_capability=THREAT_CAPABILITY,
        resistance_strength=RESISTANCE_STRENGTH,
        loss_magnitude=LOSS_MAGNITUDE,
    )

    safe_name = vendor_name.replace(" ", "_")
    xlsx_path = f"{safe_name}_control_review.xlsx"
    pdf_path  = f"{safe_name}_risk_report.pdf"

    print_report(assessment)
    export_control_review_xlsx(assessment, xlsx_path)
    export_risk_report_pdf(assessment, pdf_path)